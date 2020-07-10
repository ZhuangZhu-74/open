#coding=utf-8

import sys, os, shutil
import openpyxl
from openpyxl.utils.cell import get_column_letter, column_index_from_string

base_path = os.getcwd()
sys.path.append(base_path)


class handle_excel():
    def __init__(self, xlsx_file):
        try:
            self.file = xlsx_file
            self.wb = self.load_excel()
            self.ws = None
            
        except Exception as e:
            print('init error: ')
            print(e)
            
    @staticmethod
    def use_backup(newname, old_file=None):
        '''
        backup file
        '''
        if old_file == None:
            old_file = "c:\test.xlsx"
        if os.path.isfile(newname) and os.path.getsize(newname):
            pass
        else:
            os.path.isfile(old_file)
            shutil.copy(old_file, newname)
            
        '''
        global 分开写，不然会报错
        '''
        global excel_file
        excel_file = newname
        
        return excel_file
    
    
    def load_excel(self):
        """
        load excel file
        """
        try:
            wb = openpyxl.load_workbook(self.file)
        except Exception as e:
            print(self.file)
            print(e)
            return None
        return wb
            
    def sheet_by_name(self, name=None):
        '''
        same method name like 'xlrd'
        '''
        
        all_sheet_names = self.wb.sheetnames
        if name != None:
            if name in all_sheet_names:
                self.ws = self.wb[name]
            else:
                print("No such name {}".format(name))
                return None
        else:
            self.ws = self.wb.active
        return self.ws
    
    def sheet_by_index(self, index=None):
        '''
        same method name like 'xlrd'
        index >= 0
        '''
        
        all_sheet_names = self.wb.sheetnames
        if index in range(len(all_sheet_names)):
            self.ws = self.wb[all_sheet_names[index]]
        elif index == None:
            self.ws = self.wb.active
        else:
            print("index illegal")
            return None
        return self.ws

    def save_wb(self, excel_file=None):
        if excel_file == None:
            excel_file = self.file
        else:
            excel_file = excel_file
        self.wb.save(excel_file)
        
    @staticmethod
    def col_conv(col):
        '''
        give str or int of a sheet col
        return (col_index, col_tag) 
        '''
        if isinstance(col, str):
            col_tag = col
            col_index = column_index_from_string(col)
        elif isinstance(col, int):
            col_index = col
            col_tag = get_column_letter(col)
        return col_index, col_tag
    
    @property
    def howmany_rows(self):
        return self.ws.max_row - self.ws.min_row + 1
    
    @property
    def howmany_cols(self):
        return self.ws.max_column - self.ws.min_column + 1
        
    
    def get_lineno_by_caseid(self, caseid):
        # add 1 for 1-based excel cell handle
        first_col = self.get_col_values(1)
        #print(first_col)
        if caseid in first_col:
            return first_col.index(caseid) + 1
        else:
            return None
        
    def get_all_values_by_rows(self, exclude_firstline=False):
        """
        get all data by list of rows in sheet.
        
        ws.values
        return a generator
        """
        if exclude_firstline == False:
            return self.ws.values
        else:
            data = list(self.ws.values)
            del data[0]
            return iter(data)
        
    def get_all_values_by_cols(self):
        """
        get all data by list of cols in sheet.
        
        return a generator
        """
        
        for col in self.ws.iter_cols(values_only=True):
            yield col
    
    def get_row_values(self, row):
        """
        get one-row value, return a list
        row should start from 1
        """

        if row in range(self.ws.min_row, self.ws.max_row+1):
            return list(map(lambda i:i.value, self.ws[row]))
        else:
            return None
    
    def get_col_values(self, col):
        """
        get one-col value, return a list.
        
        if you want to get col B
        write get_col_values('B') or get_col_values(2)
        """
        
        col_index, col_tag = self.col_conv(col)
        
        if col_index in range(self.ws.min_column, self.ws.max_column+1):
            return list(map(lambda i:i.value, self.ws[col_tag]))
        else:
            return None
    
    def get_cell_value(self, row, col):
        """
        get cell value
        """
        
        col_index = self.col_conv(col)[0]
        
        if row in range(self.ws.min_row, self.ws.max_row+1):
            if col_index in range(self.ws.min_column, self.ws.max_column+1):
                return self.ws.cell(row, col_index).value

    def write_cell_value(self, row, col, value):
        """
        write a cell. but not save.
        """
        
        col_index = self.col_conv(col)[0]
        
        self.ws.cell(row, col_index, value)
        #self.wb.save(excel_file)

    def write_col_values(self, data, col, row_start=1):
        '''
        data: tuple or list for write, 
        col_tag: 
        row_start: start index of row, default is 1
        '''

        col_index = self.col_conv(col)[0]

        for i in range(len(data)):
            self.ws.cell(row_start+i, col_index, data[i])
    
    def write_row_values(self, data, row, col_start=1):
        '''
        data: tuple or list for write, 
        col_tag: 
        row_start: start index of row, default is 1
        '''
        
        col_index = col_start

        for i in range(len(data)):
            self.ws.cell(row, col_index+i, data[i])
            


if __name__ == '__main__':

    handle = handle_excel(r"C:\Users\ylliu\Desktop\API_test\test.xlsx")
    handle.sheet_by_index()
    #print(handle.ws)
    '''
    ws2 = handle.sheet_by_index()
    print(handle.ws)
    print(ws2)
    '''
    #print(handle.get_lineno_by_caseid('simu_4'))
    test = handle.get_all_values_by_rows(exclude_firstline=True)
    print(list(test))
    #handle.write_cell_value(4, 14, "test")
    #handle.save_wb()
    
    
